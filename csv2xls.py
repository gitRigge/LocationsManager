#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
#
# Copyright (c) 2023, Roland Rickborn (r_2@gmx.net)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import argparse
import csv
import datetime as dt
import glob
import locale
import os
import sys
from os.path import exists

import openpyxl

__author__ = 'Roland Rickborn'
__copyright__ = 'Copyright (c) 2024 {}'.format(__author__)
__version__ = '1.0'
__url__ = 'https://github.com/gitRigge/LocationsManager'
__license__ = 'MIT License (MIT)'


def get_save_filename(suggested_filename: str):
    retval = ''
    filename_exists = True
    counter = 1
    _sfname = suggested_filename.split('.')[0]
    _sfext = suggested_filename.split('.')[1]
    while filename_exists:
        if not exists(suggested_filename):
            retval = suggested_filename
            filename_exists = exists(suggested_filename)
        elif not exists('{}_({}).{}'.format(_sfname, counter, _sfext)):
            retval = '{}_({}).{}'.format(_sfname, counter, _sfext)
            filename_exists = exists('{}_({}).{}'.format(
                _sfname, counter, _sfext))
        counter += 1
    return retval


def get_most_possible_file():
    retval = ''
    types = ('*.csv')
    files_grabbed = []
    files_grabbed.extend(glob.glob(types))
    if files_grabbed != []:
        retval = max(files_grabbed, key=os.path.getctime)
    return retval


def write_init_output_file(outputFilename: str, outputColumns: list):
    outputFilename = get_save_filename(outputFilename)
    retval = outputFilename
    with open(outputFilename, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        outputColumns[0] = u'\uFEFF' + outputColumns[0]
        csvwriter.writerow(outputColumns)
    return retval


def append_to_output_file(outputfile: str, data: list):
    with open(outputfile, 'a', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        new_line = []
        for item in data:
            if data[item] is not None:
                new_line.append(str(data[item]).replace('\n', ''))
            else:
                new_line.append('')
        csvwriter.writerow(new_line)


def read_input_file(input_filename: str, columns: list):
    retval = {}
    header_row_keys = columns
    header_checked = False
    try:
        with open(input_filename, newline='', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            for row in csv_reader:
                if not header_checked:
                    if validate_header(row, columns):
                        header_checked = True
                    else:
                        print('Header of CSV file not correct')
                        sys.exit(12)
                t = {}
                id = row[-1]
                for col in range(0, len(row)):
                    t[header_row_keys[col]] = row[col]
                retval[id] = t
        return retval
    except FileNotFoundError as e:
        print(e)
        sys.exit(2)


def validate_header(header_row: list, columns: list):
    for i in range(0, len(columns)):
        if not header_row[i].lower().endswith(columns[i].lower().strip()):
            return False
    return True


def get_date_by_str(datetimestr: str):
    retval = ''
    if 'T' in datetimestr and '+' in datetimestr:
        try:
            d = dt.datetime.strptime(datetimestr+'00', '%Y-%m-%dT%H:%M:%S%z')
            retval = d.replace(tzinfo=None)
        except Exception:
            retval = datetimestr
    else:
        try:
            d = dt.datetime.strptime(datetimestr, '%m/%d/%Y')
            retval = d.replace(tzinfo=None)
        except Exception:
            retval = datetimestr
    return retval


def get_date_format_by_str(datetimestr: str):
    retval = ''
    loc = locale.getlocale()
    if 'T' in datetimestr and '+' in datetimestr:
        if loc[0].startswith('de'):
            retval = 'dd.mm.yyyy HH:MM:SS'
        else:
            retval = 'mm/dd/yyyy HH:MM:SS'
    elif '/' in datetimestr:
        if loc[0].startswith('de'):
            retval = 'dd.mm.yyyy'
        else:
            retval = 'mm/dd/yyyy'
    return retval


def convert_csv_to_excel(filename: str):
    my_columns = [
        'Name',
        'Address Line 1',
        'Address Line 2',
        'City',
        'Address State',
        'Zip Code',
        'Country',
        'Full Address',
        'Latitude',
        'Longitude',
        'Keywords',
        'Reserved Keywords',
        'State',
        'Last Modified',
        'Last Modified By',
        'Id'
    ]
    my_input_data = read_input_file(
        '{}.csv'.format(filename), my_columns)
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = filename.split('\\')[-1]
    my_keys = list(my_input_data.keys())
    row = 0
    for k in my_keys:
        column = 0
        cell_number = row + 1
        item = my_input_data[k]
        for c in my_columns:
            char_int = 65 + column
            cell_char = chr(char_int)
            if cell_number == 1:  # Write header line
                ws['{}{}'.format(cell_char, cell_number)] = item[c]
                ws['{}{}'.format(
                    cell_char, cell_number)].font = openpyxl.styles.Font(
                        bold=True)
            elif cell_char in ['I', 'J']:  # Write integer as text
                ws['{}{}'.format(
                    cell_char, cell_number)] = str(item[c])
            elif cell_char in ['N']:  # Detect datetime objects
                my_date = get_date_by_str(item[c])
                ws['{}{}'.format(cell_char, cell_number)] = my_date
                if isinstance(my_date, dt.datetime):
                    my_date_format = get_date_format_by_str(item[c])
                    ws['{}{}'.format(
                        cell_char, cell_number)].number_format = my_date_format
            else:
                ws['{}{}'.format(
                    cell_char, cell_number)] = item[c]
            column += 1
        row += 1
    ws.auto_filter.ref = ws.dimensions
    new_filename = get_save_filename('{}.xlsx'.format(filename))
    wb.save(new_filename)
    return new_filename


def run_from_command_line(args):
    if args.inputfile is None:
        candidate = get_most_possible_file()
        if candidate != '':
            user_input = input(
                'Do you want to continue with \'{}\' (yes/no): '.format(
                    candidate))
            if user_input.lower() == 'yes' or user_input.lower() == 'y':
                filename = '{}'.format(candidate).split('.')[0]
                output = convert_csv_to_excel(filename)
            else:
                user_input = input('Enter the filename to read: ')
                if user_input.endswith('.csv'):
                    filename = '{}'.format(user_input).split('.')[0]
                    output = convert_csv_to_excel(filename)
                else:
                    print('Wrong file format - exit')
                    sys.exit(13)
        else:
            print('Run \'csv2xls --help\' to get help')
            sys.exit(0)
    elif args.inputfile.endswith('.csv'):
        filename = '{}'.format(args.inputfile).split('.')[0]
        output = convert_csv_to_excel(filename)
    print('Output file: {}'.format(output))


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Manage SharePoint Enterprise Bookmarks',
        epilog='''For more help, see:
        {}'''.format(__url__))
    parser.add_argument(
        '-i', '--inputfile', action='store', type=str,
        help='Specify input file to read (CSV)')
    args = parser.parse_args(argv)
    run_from_command_line(args)


if __name__ == "__main__":
    main()
