#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
#
# Copyright (c) 2023, Roland Rickborn (r_2@gmx.net)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import argparse
import csv
import glob
import math
import os
import sys
from os.path import exists

import openpyxl

__author__ = 'Roland Rickborn'
__copyright__ = 'Copyright (c) 2024 {}'.format(__author__)
__version__ = '1.0'
__url__ = 'https://github.com/gitRigge/LocationsManager'
__license__ = 'MIT License (MIT)'


def get_save_filename(suggested_filename: str):
    retval = ''
    filename_exists = True
    counter = 1
    _sfname = suggested_filename.split('.')[0]
    _sfext = suggested_filename.split('.')[1]
    while filename_exists:
        if not exists(suggested_filename):
            retval = suggested_filename
            filename_exists = exists(suggested_filename)
        elif not exists('{}_({}).{}'.format(_sfname, counter, _sfext)):
            retval = '{}_({}).{}'.format(_sfname, counter, _sfext)
            filename_exists = exists('{}_({}).{}'.format(
                _sfname, counter, _sfext))
        counter += 1
    return retval


def get_most_possible_file():
    retval = ''
    types = ('*.xlsx')
    files_grabbed = []
    files_grabbed.extend(glob.glob(types))
    if files_grabbed != []:
        retval = max(files_grabbed, key=os.path.getctime)
    return retval


def write_init_output_file(outputFilename: str, outputColumns: list):
    outputFilename = get_save_filename(outputFilename)
    retval = outputFilename
    with open(outputFilename, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        outputColumns[0] = u'\uFEFF' + outputColumns[0]
        csvwriter.writerow(outputColumns)
    return retval


def append_to_output_file(outputfile: str, data: list):
    with open(outputfile, 'a', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        new_line = []
        for item in data:
            if data[item] is not None:
                new_line.append(str(data[item]).replace('\n', ''))
            else:
                new_line.append('')
        csvwriter.writerow(new_line)


def read_input_file(filename: str, columns: list):
    retval = {}
    try:
        wb = openpyxl.load_workbook('{}.xlsx'.format(filename))
        ws = wb.active
        for r in range(2, ws.max_row+1):
            retval[r] = {}
            i = 0  # 0
            retval[r][columns[i]] = ws['{}{}'.format('A', r)].value
            i += 1  # 1
            retval[r][columns[i]] = ws['{}{}'.format('B', r)].value
            i += 1  # 2
            retval[r][columns[i]] = ws['{}{}'.format('C', r)].value
            i += 1  # 3
            retval[r][columns[i]] = ws['{}{}'.format('D', r)].value
            i += 1  # 4
            retval[r][columns[i]] = ws['{}{}'.format('E', r)].value
            i += 1  # 5
            retval[r][columns[i]] = ws['{}{}'.format('F', r)].value
            i += 1  # 6
            retval[r][columns[i]] = ws['{}{}'.format('G', r)].value
            i += 1  # 7
            retval[r][columns[i]] = ws['{}{}'.format('H', r)].value
            i += 1  # 8
            retval[r][columns[i]] = ws['{}{}'.format('I', r)].value
            i += 1  # 9
            retval[r][columns[i]] = ws['{}{}'.format('J', r)].value
            i += 1  # 10
            retval[r][columns[i]] = ws['{}{}'.format('K', r)].value
            i += 1  # 11
            retval[r][columns[i]] = ws['{}{}'.format('L', r)].value
            i += 1  # 12
            retval[r][columns[i]] = ws['{}{}'.format('M', r)].value
            i += 1  # 13
            retval[r][columns[i]] = ws['{}{}'.format('N', r)].value
            i += 1  # 14
            retval[r][columns[i]] = ws['{}{}'.format('O', r)].value
            i += 1  # 15
            retval[r][columns[i]] = ws['{}{}'.format('P', r)].value
        return retval
    except FileNotFoundError as e:
        print(e)
        sys.exit(2)


def convert_excel_to_csv(filename: str):
    retval = []
    my_output_columns = [
        'Name',
        'Address Line 1',
        'Address Line 2',
        'City',
        'Address State',
        'Zip Code',
        'Country',
        'Full Address',
        'Latitude',
        'Longitude',
        'Keywords',
        'Reserved Keywords',
        'State',
        'Last Modified',
        'Last Modified By',
        'Id'
    ]
    my_content = read_input_file(filename, my_output_columns)
    limit = 3000
    counter = 0
    number_of_files = math.ceil(len(my_content)/limit)
    if number_of_files > 1:
        for i in range(0, number_of_files):
            my_output_filename = write_init_output_file('{}_{}.csv'.format(
                filename, i+1), my_output_columns)
            retval.append(my_output_filename)
            for j in range(0, limit):
                my_keys = list(my_content)
                if counter < len(my_keys):
                    my_bm = my_content[my_keys[counter]]
                    append_to_output_file(
                        my_output_filename, my_bm)
                    counter += 1
    else:
        my_output_filename = write_init_output_file('{}.csv'.format(
            filename), my_output_columns)
        retval.append(my_output_filename)
        for j in range(0, limit):
            my_keys = list(my_content)
            if counter < len(my_keys):
                my_bm = my_content[my_keys[counter]]
                append_to_output_file(my_output_filename, my_bm)
                counter += 1
    return ', '.join(retval)


def run_from_command_line(args):
    if args.inputfile is None:
        candidate = get_most_possible_file()
        if candidate != '':
            user_input = input(
                'Do you want to continue with \'{}\' (yes/no): '.format(
                    candidate))
            if user_input.lower() == 'yes' or user_input.lower() == 'y':
                filename = '{}'.format(candidate).split('.')[0]
                output = convert_excel_to_csv(filename)
            else:
                user_input = input('Enter the filename to read: ')
                if user_input.endswith('.xlsx'):
                    filename = '{}'.format(user_input).split('.')[0]
                    output = convert_excel_to_csv(filename)
                else:
                    print('Wrong file format - exit')
                    sys.exit(13)
        else:
            print('Run \'xls2csv --help\' to get help')
            sys.exit(0)
    elif args.inputfile.endswith('.xlsx'):
        filename = '{}'.format(args.inputfile).split('.')[0]
        output = convert_excel_to_csv(filename)
    print('Output file: {}'.format(output))


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Manage SharePoint Enterprise Bookmarks',
        epilog='''For more help, see:
        {}'''.format(__url__))
    parser.add_argument(
        '-i', '--inputfile', action='store', type=str,
        help='Specify input file to read (Excel)')
    args = parser.parse_args(argv)
    run_from_command_line(args)


if __name__ == "__main__":
    main()
